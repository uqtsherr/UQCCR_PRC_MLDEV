from __future__ import division, print_function, absolute_import
import tensorflow as tf
from tensorflow.keras import Model, layers
from tensorflow.keras.mixed_precision import experimental as mixed_precision
#policy = mixed_precision.Policy('mixed_float16')
#mixed_precision.set_policy(policy)
import h5py
import numpy as np
#import matplotlib.pyplot as plt
import os.path
from os import listdir
from os.path import isfile, join, isdir, splitext
from random import uniform
import datetime
import pyedflib
import openpyxl
version_maj = 1
version_min = 2
print('piglet RNN v%i.%i'% (version_maj,version_min))

foldername = '/scratch/medicine/TS-PRC/source_data/'
fileList = [join(foldername, f) for f in listdir(foldername) if isfile(join(foldername, f))]

fs =256
channels = 3
signal_labels = []
signal_nsamples = []
Records = []
for f in fileList:
    a = splitext(f)
    if a[1] == '.mat':
        with h5py.File(f) as file:
            print(file['Records'].keys())
            Class = file['Records']['ClassNum'].value
            print(Class)
            xTFDs = file['Records']['TFDs'][()]
            xEEG = file['Records']['eeg']
            xFeatures = file['Records']['features']

catDict = {'Normal': 0, 'Burst': 1, 'Suppression': 2, 'Artifact': 3, 'Seizure': 4}
num_classes = 5
if isfile(join(foldername, 'EEG_annotation_recovered.csv.xlsx')):
    annotationsWB = openpyxl.load_workbook(join(foldername, 'EEG_annotation_recovered.csv.xlsx'), read_only=True)
print(annotationsWB.sheetnames)
eegdata = []
cnts = []
for file in fileList:
    fpath = join(foldername,file)
    p, ext = os.path.splitext(fpath)
    if(ext == '.EDF'):
        print(fpath + " is a .edf file")
        annotation = list(filter(lambda a: a in p, annotationsWB.sheetnames))
        if annotation!=[]:
            print("file contains a record :)")
            edf = pyedflib.EdfReader(fpath)
            samples = edf.getNSamples()
            dat = np.zeros([samples[0],channels])
            for ch in range(channels):
                dat[:,ch] = edf.readSignal(ch, start=0, n=samples[ch])
            i = 0
            print("length in seconds: %f" % (samples[0]/fs))
            ws = annotationsWB[annotation[0]]  # get sheet for curr rec
            labels = np.zeros([int(samples[0]/fs),num_classes+1])
            for row in ws.iter_rows(min_row=4, max_col=2, max_row=int(samples[0]/fs+2),
                                    values_only=True):
                labels[i, catDict.get(row[1], 5)] = 1
                i += 1
            labels = labels[0:i-9,:]
            cnts =+ np.sum(labels, 0)
            print("length annotations: %f" % (i-9))
            print(labels)
            eegdata.append((samples,edf,annotation[0],dat,labels))
            #nSigs = edf.signals_in_file
            #sig1 = np.zeros((nSigs, edf.getNSamples()[0]))
            #for i in np.arange(nSigs):
            #    sig1[i, :edf.getNSamples()[i]] = edf.readSignal(i)
            #datName = (sig1,p)
        else:
            print("edf file was not annotated :(")
    else:
        print(fpath + " is not a .edf file")
    #break #just grab 1 file - for testing speedup

print(cnts)
totalSamples = 0;
for record in eegdata:
    totalSamples += record[0]
prob = []
for record in eegdata:
    p = record[0]/totalSamples
    prob.append(p[0])
records = (eegdata, annotationsWB,totalSamples[0])
fs = records[0][0][1].getSampleFrequency(0)
#assumes 1 second classification sectors
def get_batch(records , prob , n_seconds_per_seq , n_sequences):
    block = np.zeros([n_sequences,n_seconds_per_seq, channels, 2 * fs])
    labels = np.zeros([n_sequences,n_seconds_per_seq,num_classes+1])
    for k in range(n_sequences):
        p = uniform(0, 1)
        currRec = []
        # find which record we are using
        for i in range(len(prob)):
            p -= prob[i]
            if(p<0):
                curr_rec = records[0][i]
                break
        start = int(uniform(0,(curr_rec[4].shape[0]-n_seconds_per_seq * n_sequences)*fs)) #randomly pick a start point across the record that will not over/underrun
        tmp = np.zeros([1,(n_seconds_per_seq+1)*fs])
        for c in range(channels):
            #tmp = curr_rec[1].readSignal(c, start=start, n=(n_seconds_per_seq+1)*fs)
            tmp = curr_rec[3][start:start+(n_seconds_per_seq+1)*fs,c]
            for i in range(n_seconds_per_seq):
                block[k,i,c,:] = tmp[fs*i:fs*(i+2)]
        offset = int(start/fs)

        #ws = records[1][curr_rec[2]] #get sheet for curr rec
        #catDict = {'Normal':0,'Burst':1,'Suppression':2,'Artifact':3,'Seizure':4}
        #i = 0
        #for row in ws.iter_rows(min_row=offset+1, max_col=2, max_row=offset+n_seconds_per_seq, values_only=True):
        #    labels[k,i,catDict.get(row[1], -1)] =  1
        #    i += 1
        labels[k,:,:] = curr_rec[4][offset:offset+n_seconds_per_seq,:]
    return block, labels

def get_dataset(records , n_seconds_per_seq):
    n_seq = 0
    for c in records[0]:
        n_seq +=  int(c[4].shape[0]/n_seconds_per_seq)-1
    block = np.zeros([n_seq, n_seconds_per_seq, channels, 2 * fs])
    labels = np.zeros([n_seq, n_seconds_per_seq, num_classes+1])
    #print(np.shape(block))
    #print(np.shape(labels))
    seq_start = 0 #tracks start of sequences between records
    for curr_rec in records[0]:
        #print(seq_start)
        n_sequences = int(curr_rec[4].shape[0]/n_seconds_per_seq)-1
        #n_sequences = int(curr_rec[1].getNSamples()[0]/fs/n_seconds_per_seq)-1
        #block = np.zeros([n_sequences,n_seconds_per_seq, channels, 2 * fs])
        #labels = np.zeros([n_sequences,n_seconds_per_seq,5])
        seq_offset = int(uniform(0,n_seconds_per_seq)) #randomly start the sequence within the sequence gap range
        for k in range(n_sequences):
            start = int(uniform(0,256)+fs*(k*n_seconds_per_seq+seq_offset)) #randomly wiggle the start point of a sequence across 1second
            tmp = np.zeros([1,(n_seconds_per_seq+1)*fs])
            for c in tf.range(channels):
                #tmp = curr_rec[1].readSignal(c, start=start, n=(n_seconds_per_seq+1)*fs)
                tmp = curr_rec[3][start:start+(n_seconds_per_seq+2)*fs,c]
                for i in range(n_seconds_per_seq):
                    block[k+seq_start,i,c,:] = tmp[fs*i:fs*(i+2)]
            offset = int(start/fs)
            labels[k+seq_start,:,:] = curr_rec[4][offset:offset+n_seconds_per_seq,:]
        seq_start += n_sequences

    return block, labels

##End data load section - records is a tuple of edfFiles,sheet string
#Data is currently imported as a record, where records are made up of a (label data) tuple. label file of N x 3 samples where N is record length in seconds
#The a data is matrix of 21x (NxFs) where fs is 256Hz

# Training Parameters
learning_rate = 0.00002
training_epochs = 150
batch_size = 10
sequence_length = 75
save_epoch = 3

train_CNN = True
dropout = 0.25
LSTM_output_units = 200
LSTM_input_units = 200

#print('testing load times')
#bl, lb = get_batch(records,prob,sequence_length,batch_size)
#print(np.shape(bl))
#print(np.shape(lb))

#dfefine logging paths
logs_path = '/scratch/medicine/TS-PRC/piglet/logs/v%i.%i/'%(version_maj,version_min)
modelPath = '/scratch/medicine/TS-PRC/piglet/v%i.%i/checkpoint/model.ckpt'%(version_maj,version_min)
print('saving to:'+ modelPath)
current_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
train_summary_writer = tf.summary.create_file_writer(logs_path + current_time + '/train')
test_summary_writer = tf.summary.create_file_writer(logs_path + current_time + '/test')




class LSTM(Model):
    # Set layers.
    def __init__(self):
        super(LSTM, self).__init__()
        # Define a Masking Layer with -1 as mask.
        self.norm = x = layers.BatchNormalization()
        self.CNN = tf.keras.applications.EfficientNetB0(
            include_top=False, weights='imagenet', input_tensor=None,
            input_shape=None, pooling=None, classes=1000)
        self.CNN.trainable = train_CNN
        self.dense = layers.Dense(LSTM_input_units)
        # Define a LSTM layer to be applied over the Masking layer.
        # Dynamic computation will automatically be performed to ignore -1 values.
        self.lstm = layers.Bidirectional(layers.GRU(units=LSTM_output_units, return_sequences=True,dropout=dropout))
        # Output fully connected layer (5 classes).
        self.out = layers.Dense(num_classes+1)


    # Set forward pass.
    #input dims:[batch_size, seq_len, channels, window]
    def call(self, x, is_training=False):
        #turn EEG from temporal signals to spectrogram imgs
        x = tf.signal.stft(x,128,4,window_fn=tf.signal.hann_window)
        #x = tf.math.log(x)
        x = self.norm(x)
        x = tf.transpose(x,perm=[0,1,3,4,2])
        x = tf.reshape(x, shape=[-1,97,65,3])
        x = self.CNN(x)
        x = tf.reshape(x, shape=[-1,3*2*1280])
        x = self.dense(x)
        # A RNN Layer expects a 3-dim input (batch_size, seq_len, num_features).
        x = tf.reshape(x, shape=[-1, sequence_length, LSTM_input_units])
        # No Masking layer as I have set seq length.
        # Apply LSTM layer.
        x = self.lstm(x)
        # Apply output layer.
        x = self.out(x)
        if not is_training:
            # tf cross entropy expect logits without softmax, so only
            # apply softmax when not training.
            x = tf.nn.softmax(x)
        return x

LSTM_net = LSTM()
LSTM_net.build([batch_size,sequence_length,channels,fs*2])
LSTM_net.summary()



# Create a callback that saves the model's weights
ckpt = tf.train.Checkpoint(LSTM_net)



latest = tf.train.latest_checkpoint(modelPath)
print(latest)
# Restore the model
load_result = ckpt.restore(latest)

try:
    print(load_result)
    load_result.assert_existing_objects_matched()
    print('loaded model')
except:
    print('model not loaded')

# Cross-Entropy Loss.
# Note that this will apply 'softmax' to the logits.
def cross_entropy_loss(x, y):
    # Convert labels to int 64 for tf cross-entropy function.
    y = tf.cast(y, tf.int64)
    # Apply softmax to logits and compute cross-entropy.
    loss = tf.nn.softmax_cross_entropy_with_logits(labels=y, logits=x) #sparse_softmax_cross_entropy_with_logits
    # Average loss across the batch.
    return tf.reduce_mean(loss)

# Accuracy metric.
def accuracy(y_pred, y_true):
    # Predicted class is the index of highest score in prediction vector (i.e. argmax).
    correct_prediction = tf.equal(tf.argmax(y_true, 2), tf.argmax(y_pred, 2))
    #print(correct_prediction)
    return tf.reduce_mean(tf.reduce_mean(tf.cast(correct_prediction, tf.float32), axis=-1))

# Adam optimizer.
optimizer = tf.optimizers.Adam(learning_rate)


# Optimization process.
def run_optimization(x, y):
    # Wrap computation inside a GradientTape for automatic differentiation.
    with tf.GradientTape() as g:
        # Forward pass.
        pred = LSTM_net(x, is_training=True)
        # Compute loss.
        loss = cross_entropy_loss(pred, y)

    # Variables to update, i.e. trainable variables.
    trainable_variables = LSTM_net.trainable_variables

    # Compute gradients.
    gradients = g.gradient(loss, trainable_variables)

    # Update weights following gradients.
    optimizer.apply_gradients(zip(gradients, trainable_variables))
    acc = accuracy(pred, y)
    return loss, acc


# Run training for the given number of steps.
for step in range(training_epochs):
    SHUFFLE_BUFFER_SIZE = 5000
    N_test = 300
    x, y = get_dataset(records, sequence_length)
    dataset = tf.data.Dataset.from_tensor_slices((x, y))
    dataset = dataset.shuffle(SHUFFLE_BUFFER_SIZE).batch(batch_size)
    test_dataset = dataset.take(N_test)
    train_dataset = dataset.skip(N_test)
    batch = 0
    train_loss = tf.convert_to_tensor(0.0)
    train_acc = tf.convert_to_tensor(0.0)
    for batch_x, batch_y in train_dataset:
        # Run the optimization to update W and b values.
        #batch_x, batch_y = get_batch(records, prob, sequence_length, batch_size)
        l, a = run_optimization(batch_x, batch_y)
        train_acc = train_acc + a
        train_loss = train_loss + l
        batch +=1
        if batch % 50 == 0:
            print("batch: %i, running average loss: %f, running average accuracy: %f" % (batch, train_loss/batch, train_acc/batch))

    with train_summary_writer.as_default():
        tf.summary.scalar('loss', train_loss/batch, step=step)
        tf.summary.scalar('accuracy', train_acc/batch, step=step)
    loss = tf.convert_to_tensor(0.0)
    acc = tf.convert_to_tensor(0.0)
    for test_x, test_y in test_dataset:
    #test_x, test_y = get_batch(records, prob, sequence_length, batch_size)
        pred = LSTM_net(test_x, is_training=True)
        loss = cross_entropy_loss(pred, test_y) + loss
        acc = accuracy(pred, test_y) + acc
    acc = acc / N_test
    loss = loss / N_test
    print("epoch: %i, test av loss: %f, av accuracy: %f" % (step, loss, acc))
    with test_summary_writer.as_default():
        tf.summary.scalar('loss', loss, step=step)
        tf.summary.scalar('accuracy', acc, step=step)

    if step % save_epoch == 0 & step != 0:
        save_path = ckpt.save(modelPath)
        print('saved model')


records[1].close() #close the xl sheet